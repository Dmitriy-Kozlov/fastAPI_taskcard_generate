import json
from openpyxl import load_workbook
import re
from collections import defaultdict
import os
from docx import Document
import subprocess

SOURCE_PATH = "files/source"


def collect_from_mpd(mpd_file, atype):
    # Загрузить Excel-файл
    workbook = load_workbook(mpd_file, data_only=True)
    sheet = workbook["MPD"]
    os.makedirs(f"{SOURCE_PATH}", exist_ok=True)

    results = {}

    # Пройти по строкам начиная со второй (первая — заголовки)
    for row in sheet.iter_rows(min_row=4):
        task_number = row[2].value  # колонка C

        if not task_number:
            continue

        preparation = row[5].value  # колонка F
        zone = row[6].value         # колонка G
        skill_code = row[8].value   # колонка I
        interval = row[13].value    # колонка N
        try:
            cleaned_lines = []
            for line in interval.splitlines():
                if "or" in line.lower() or "note" in line.lower():
                    continue
                cleaned_lines.append(line.strip())

            cleaned_interval = "\n".join(cleaned_lines)
        except:
            cleaned_interval = ""
        manhours_rows = []
        if atype == "A330":
            reference = row[26].value   # колонка AA
            manhours_rows = [28, 29, 30] # Колонки AC, AD, AE — индексы 28, 29, 30
        if atype == "A321":
            reference = row[18].value   # колонка S
            manhours_rows = [20, 21, 22] # Колонки U, V, W — индексы 20, 21, 22


        manhour_parts = []

        for i in manhours_rows:
            val = row[i].value
            subtotal = 0

            if isinstance(val, str):
                # Разделяем по строкам и суммируем каждую строку как float
                try:
                    subtotal = round(sum(float(x.strip()) for x in val.splitlines() if x.strip()), 2)

                except ValueError:
                    subtotal = 0  # если не удается преобразовать в float
            elif isinstance(val, (int, float)):
                subtotal = round(float(val), 2)

            manhour_parts.append(subtotal)

        total_manhour = round(sum(manhour_parts), 2)

        results[task_number] = {
            "preparation": preparation,
            "zone": zone,
            "skill_code": skill_code,
            "interval": cleaned_interval,
            "reference": reference,
            "manhour": total_manhour if total_manhour else "TBD"
        }

    # Сохранить результат в JSON-файл
    with open(f"{SOURCE_PATH}/{atype}_mpd_data.json", "w", encoding="utf-8") as f:
        json.dump(results, f, indent=2, ensure_ascii=False)


def parse_IPC(ipc_file, atype):

    ipc_results = {}

    with open(ipc_file, "r", encoding="utf-8") as f:
        # with open("test_IPC.json", "r", encoding="utf-8") as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                print(f"Ошибка в строке {line_num}: {e}")
                continue
            # print(obj)
            title = obj.get("title", "")
            try:
                text = obj.get("contents", [{}])[0].get("text", "")
            except IndexError:
                text = ""

            # Извлекаем IPC number (4 группы цифр через тире)
            # ipc_number_match = re.search(r'\b(\d{2}-\d{2}-\d{2}-\d{2})\b', title)
            ipc_number_match = re.search(r'\b(\d{2}-\d{2}-\d{2}-\d{2})[A-Z]?\b', title)
            ipc_number = ipc_number_match.group(1) if ipc_number_match else None

            # Извлекаем название после " - "
            if " - " in title:
                ipc_title = title.split(" - ", 1)[1]
                ipc_title = re.sub(r"\(.*?\)", "", ipc_title).strip()
            else:
                ipc_title = title.strip()

            # Инициализация словаря
            if not ipc_number:
                continue
            ipc_results[ipc_number] = {
                "ipc_title": ipc_title,
                # "text": text.strip(),
                "items": []
            }

            # Извлекаем блок после "ITEM PART NUMBER ..."
            split_text = re.split(r'ITEM PART NUMBER NOMENCLATURE.*?UNIT PER ASSY', text,
                                  flags=re.DOTALL | re.IGNORECASE)
            if len(split_text) < 2:
                continue

            items_block = split_text[1]

            # Регулярка: 3 цифры (Item), пробел, part number, пробел, точка, текст (игнорируем), и в конце 3 цифры (QTY)
            item_pattern = re.findall(
                # r'\b(\d{3})\s+([A-Z0-9\-\.]+)\s+\.\s+.*?\b(\d{3})\b',
            r'\b(\d{3})\s+(?:Installation context\s+)?([A-Z0-9\-]+)\s+\.\s+[^0-9].*?\b(\d{3})\b',
                items_block
            )

            for item, part_number, qty in item_pattern:
                ipc_results[ipc_number]["items"].append({
                    "item": item,
                    "part_number": part_number,
                    "qty": qty
                })

    # Сохраняем результат
    with open(f"{SOURCE_PATH}/{atype}_parsed_IPC_results.json", "w", encoding="utf-8") as out_file:
        json.dump(ipc_results, out_file, indent=2, ensure_ascii=False)


def parse_tool_material_from_AMM(amm_file, atype):

    def parse_tools_and_materials(text, tool_list, id_tool_list, material_list):
        tools = []
        result_tool_list = []
        result_material_list = []

        materials = []
        # Парсим секцию Fixtures, Tools, Test and Support Equipment

        fixtures_match = re.search(
            r"Fixtures, Tools, Test and Support Equipment\s*(.*?)\s*(?:[B-D]\.|Work Zones)",
            text,
            re.DOTALL | re.IGNORECASE
        )

        if fixtures_match:
            lines = fixtures_match.group(1).splitlines()[0]
            cleaned = re.sub(r'Search in TEM \(.*?\)\s*Search in GSE Cockpit ', '', lines)

            segments = re.split(r'(?=No specific)', cleaned)

            results = []
            seen = set()

            for segment in segments:
                match = re.match(r'No specific(?:\s+(AR|\d+))?\s+(.+)', segment.strip())
                if match:
                    qty = match.group(1) or 'AR'
                    designation = match.group(2).strip()

                    # Обрезаем по известному коду или следующему "No specific"
                    stop_tokens = ['No specific'] + id_tool_list
                    for token in stop_tokens:
                        if token in designation:
                            designation = designation.split(token)[0].strip()

                    # Убираем хвостовые дефисы
                    designation = re.sub(r'\s*[-–—]+\s*$', '', designation)

                    key = (designation, qty)
                    if key not in seen:
                        result_tool_list.append({
                            "REFERENCE": "No specific",
                            "QTY": qty,
                            "DESIGNATION": designation
                        })
                        seen.add(key)

            if id_tool_list:

                # Создаем шаблон с объединением всех идентификаторов
                pattern = rf"({'|'.join(map(re.escape, id_tool_list))})\s+(\S+)\s+(.*?)(?=(?:{'|'.join(map(re.escape, id_tool_list))})|\Z)"

                for match in re.finditer(pattern, cleaned):
                    reference, qty, designation = match.groups()
                    result_tool_list.append({
                        "REFERENCE": reference,
                        "QTY": qty,
                        "DESIGNATION": designation.strip()
                    })

        # Парсим секцию Consumable Materials

        materials_match = re.search(
            r"\. Consumable Materials\s+REFERENCE\s+DESIGNATION\s+(.*?)(?=\n\s*[C-D]\.|Referenced Information|ZONE/ACCESS|Ref\.|SUBTASK|\Z)",
            text,
            re.DOTALL
        )
        if materials_match:
            lines = materials_match.group(1).splitlines()[0]
            for code in material_list:
                # Ищем шаблон: код, пробел, описание, пробел, дефис, пробел
                match = re.search(rf'{code}\s+(.*?)(?=\s+-\s+)', lines)

                pattern = '|'.join(re.escape(code) for code in material_list)

                # Поиск всех кодов и их позиций
                matches = list(re.finditer(rf'\b({pattern})\b', lines))
                for i, match in enumerate(matches):
                    code = match.group(1)
                    start = match.end()

                    if i + 1 < len(matches):
                        end = matches[i + 1].start()
                    else:
                        # ищем конец по ключевому заголовку или дефису
                        next_stop = re.search(r'\s[-–—]\s|(?=\b[A-Z]\.)', lines[start:])
                        end = start + next_stop.start() if next_stop else len(lines)

                    designation = lines[start:end].strip(" -")

                    entry = {
                        "REFERENCE": code,
                        "DESIGNATION": designation.strip(),
                        "QTY": "AR"
                    }

                    # Уникальность по REFERENCE + DESIGNATION
                    if not any(
                            e["REFERENCE"] == entry["REFERENCE"] and e["DESIGNATION"] == entry["DESIGNATION"]
                            for e in result_material_list
                    ):
                        result_material_list.append(entry)

            no_specific_matches = re.finditer(r'No specific\s+(.*?)(?=\s+No specific|\s+[A-Z]\.|\s*$)', lines)
            for match in no_specific_matches:
                designation = match.group(1).strip()
                if designation:
                    result_material_list.append({
                        "REFERENCE": "No specific",
                        "DESIGNATION": designation,
                        "QTY": "AR"
                    })
        match = re.search(r'[A-Z]\. Expendable Parts(.*?)(?:\n\s*[D-Z]\.|SUBTASK|\*\* ON A/C|\Z)', text, re.DOTALL)
        if match:
            expendable_block = match.group(1)
            # 2. Найдём все шаблоны вида 49-81-45-01 ITEM 125
            # Ограничим блок до 'D.' включительно
            block_match = re.search(r'(.*?)\b[A-Z]\.', expendable_block, re.DOTALL)
            # print(block_match)
            if block_match:
                expendable = block_match.group(1)
                # Регулярка: находим designation, IPC, ITEM
                pattern = re.findall(
                    r'\b\d+\s+([A-Z][A-Z0-9 ,\-]+?)\s+(\d{2}-\d{2}-\d{2}-\d{2})\s+ITEM\s+(\d{3})',
                    expendable
                )

                # Формируем результат
                results = []
                for designation, ipc, item in pattern:
                    results.append({
                        "ipc_number": ipc,
                        "item": item,
                        "designation": designation.strip()
                    })

                # 3. Результат как список словарей
                expendable_refs = [{"designation": designation, "ipc_number": ipc, "item": item} for
                                   designation, ipc, item in pattern]

                with open(f"{SOURCE_PATH}/{atype}_parsed_IPC_results.json", "r", encoding="utf-8") as f:
                    ipc_data = json.load(f)

                # Поиск part_number и qty
                for ref in expendable_refs:
                    ipc_number = ref["ipc_number"]
                    item_code = ref["item"]

                    if ipc_number in ipc_data:
                        item_list = ipc_data[ipc_number].get("items", [])
                        found_item = next((i for i in item_list if i["item"] == item_code), None)

                        if found_item:
                            result_material_list.append({
                                "REFERENCE": found_item["part_number"],
                                "DESIGNATION": ref["designation"],
                                "QTY": int(found_item["qty"])
                            })

        return result_tool_list, result_material_list

    results_by_mpd = {}

    with open(amm_file, "r", encoding="utf-8") as f:
        for line_num, line in enumerate(f, 1):
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except json.JSONDecodeError as e:
                print(f"Ошибка в строке {line_num}: {e}")
                continue

            title = obj.get("title", "")
            if " - " in title:
                task_number, description = title.split(" - ", 1)
            else:
                task_number = ""
                description = title

            # Словари ключевых слов
            keywords = defaultdict(list)
            for kw in obj.get("keywords", []):
                kw_type = kw.get("type")
                kw_value = kw.get("value")
                if kw_type in {
                    "MPDREFERENCE_KEYWORD",
                    "STANDARDTOOL_KEYWORD",
                    "TOOLIDENTIFIER_KEYWORD",
                    "MATERIALCODE_KEYWORD",
                }:
                    keywords[kw_type].append(kw_value)

            mpd_key = keywords["MPDREFERENCE_KEYWORD"][0] if keywords["MPDREFERENCE_KEYWORD"] else None
            if not mpd_key:
                continue

            # Получаем текст
            content_text = ""
            for c in obj.get("contents", []):
                if "text" in c:
                    content_text = c["text"]
                    break
            # Парсим таблицы из текста
            tools_section, materials_section = parse_tools_and_materials(content_text, keywords["STANDARDTOOL_KEYWORD"],
                                                                         keywords["TOOLIDENTIFIER_KEYWORD"],
                                                                         keywords["MATERIALCODE_KEYWORD"])
            # Собираем расширенные записи
            tools_detailed = []
            for name in keywords["STANDARDTOOL_KEYWORD"]:
                match = next((t for t in tools_section if t["DESIGNATION"].strip() == name.strip()), None)
                if match:
                    tools_detailed.append({
                        "REFERENCE": match["REFERENCE"],
                        "DESIGNATION": match["DESIGNATION"],
                        "QTY": match["QTY"]
                    })

            tool_ids_detailed = []
            for ref in keywords["TOOLIDENTIFIER_KEYWORD"]:
                match = next((t for t in tools_section if t["REFERENCE"].strip() == ref.strip()), None)
                if match:
                    tool_ids_detailed.append({
                        "REFERENCE": match["REFERENCE"],
                        "QTY": match["QTY"],
                        "DESIGNATION": match["DESIGNATION"]
                    })

            materials_detailed = []
            for ref in keywords["MATERIALCODE_KEYWORD"]:
                match = next((m for m in materials_section if m["REFERENCE"].strip() == ref.strip()), None)
                if match:
                    materials_detailed.append(match)
            # Финальный результат
            results_by_mpd[mpd_key] = {
                "task_number": task_number,
                "description": description,
                "STANDARDTOOL_KEYWORD": tools_section,
                "MATERIALCODE_KEYWORD": materials_section
            }

    # Сохраняем результат
    with open(f"{SOURCE_PATH}/{atype}_parsed_results_from_AMM.json", "w", encoding="utf-8") as out_file:
        json.dump(results_by_mpd, out_file, indent=2, ensure_ascii=False)


def merge_AMM_MPD(atype):

    # Загрузка первого JSON-файла (куда нужно добавить данные)
    with open(f"{SOURCE_PATH}/{atype}_mpd_data.json", "r", encoding="utf-8") as f:
        mpd_data = json.load(f)

    # Загрузка второго JSON-файла (откуда берутся tools/materials)
    with open(f"{SOURCE_PATH}/{atype}_parsed_results_from_AMM.json", "r", encoding="utf-8") as f:
        tech_data = json.load(f)

    # Объединение данных на основе первых 6 символов названия задачи
    for mpd_task, mpd_content in mpd_data.items():
        base_key = mpd_task[:9]
        matched_key = None

        # Поиск соответствия в техническом JSON
        for tech_task in tech_data.keys():
            if tech_task.startswith(base_key):
                matched_key = tech_task
                break

        if matched_key:
            tech_info = tech_data[matched_key]

            # Копируем ключи, если они есть
            for key in ["task_number", "description", "STANDARDTOOL_KEYWORD", "TOOLIDENTIFIER_KEYWORD",
                        "MATERIALCODE_KEYWORD"]:
                if key in tech_info:
                    mpd_content[key] = tech_info[key]

    # Сохранение результата в новый JSON-файл
    with open(f"{SOURCE_PATH}/{atype}_merged_result.json", "w", encoding="utf-8") as f:
        json.dump(mpd_data, f, indent=2, ensure_ascii=False)

